import os
import re
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, field


DATE_PATTERNS = [
    r"(\d{4})[-_](\d{1,2})[-_](\d{1,2})",
    r"(\d{4})(\d{2})(\d{2})",
]


@dataclass
class ArchiveEntry:
    project_name: str
    date: str
    check_report_txt: bool = False
    check_report_csv: bool = False
    check_report_xlsx: bool = False
    handover_package: bool = False
    feedback_report_txt: bool = False
    feedback_report_xlsx: bool = False
    history_report: bool = False
    note: str = ""

    @property
    def has_check_report(self) -> bool:
        return self.check_report_txt or self.check_report_csv or self.check_report_xlsx

    @property
    def has_feedback_report(self) -> bool:
        return self.feedback_report_txt or self.feedback_report_xlsx

    @property
    def completeness(self) -> float:
        total = 4
        done = 0
        if self.has_check_report:
            done += 1
        if self.handover_package:
            done += 1
        if self.has_feedback_report:
            done += 1
        if self.history_report:
            done += 1
        return done / total if total > 0 else 0


def _extract_date_from_filename(filename: str) -> Optional[str]:
    for pat in DATE_PATTERNS:
        m = re.search(pat, filename)
        if m:
            try:
                y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
                return datetime(y, mo, d).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def _extract_project_and_date(filename: str) -> Tuple[Optional[str], Optional[str]]:
    date_str = _extract_date_from_filename(filename)
    if not date_str:
        return None, None

    pat = re.compile(r"(.+?)[-_]\d{4}[-_]?\d{1,2}[-_]?\d{1,2}[-_]")
    m = pat.match(filename)
    if m:
        return m.group(1), date_str

    return None, date_str


def scan_archive(report_dir: str) -> Dict[str, Dict[str, ArchiveEntry]]:
    archive: Dict[str, Dict[str, ArchiveEntry]] = {}

    if not os.path.isdir(report_dir):
        return archive

    for name in os.listdir(report_dir):
        path = os.path.join(report_dir, name)

        if os.path.isdir(path) and name.endswith("_交接包"):
            proj_name, date_str = _extract_project_and_date(name)
            if proj_name and date_str:
                archive.setdefault(proj_name, {}).setdefault(
                    date_str, ArchiveEntry(project_name=proj_name, date=date_str)
                ).handover_package = True
            continue

        if not os.path.isfile(path):
            continue

        proj_name, date_str = _extract_project_and_date(name)
        if not proj_name or not date_str:
            continue

        entry = archive.setdefault(proj_name, {}).setdefault(
            date_str, ArchiveEntry(project_name=proj_name, date=date_str)
        )

        if name.endswith("_桩基校核报告.txt"):
            entry.check_report_txt = True
        elif name.endswith("_桩基校核报告.csv"):
            entry.check_report_csv = True
        elif name.endswith("_桩基校核报告.xlsx"):
            entry.check_report_xlsx = True
        elif name.endswith("_整改跟踪报告.txt"):
            entry.feedback_report_txt = True
        elif name.endswith("_整改跟踪报告.xlsx"):
            entry.feedback_report_xlsx = True
        elif "_多日复查." in name or "_to_" in name and ("_多日复查" in name):
            entry.history_report = True

    return archive


def generate_archive_report_text(report_dir: str,
                                 archive: Dict[str, Dict[str, ArchiveEntry]],
                                 project_name: Optional[str] = None) -> str:
    lines = []
    lines.append("=" * 78)
    lines.append("  桩基施工记录校核 — 归档台账")
    lines.append("=" * 78)
    lines.append(f"  报告目录: {report_dir}")
    lines.append(f"  生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    if project_name:
        lines.append(f"  项目过滤: {project_name}")
    lines.append("-" * 78)
    lines.append("")

    if project_name:
        projects = [project_name] if project_name in archive else []
    else:
        projects = sorted(archive.keys())

    for proj in projects:
        dates = sorted(archive[proj].keys())
        if not dates:
            continue

        lines.append(f"【{proj}】 共 {len(dates)} 天资料")
        lines.append("-" * 78)
        lines.append(
            f"  {'日期':<12}  {'校核':<4}  {'交接包':<4}  {'整改':<4}  "
            f"{'复查':<4}  {'完整度':<8}  备注"
        )
        lines.append(f"  {'─' * 12}  {'─'*4}  {'─'*4}  {'─'*4}  {'─'*4}  {'─'*8}  {'─'*20}")

        for d in dates:
            e = archive[proj][d]
            check_icon = "✓" if e.has_check_report else "✗"
            handover_icon = "✓" if e.handover_package else "✗"
            fb_icon = "✓" if e.has_feedback_report else "✗"
            hist_icon = "✓" if e.history_report else "✗"
            comp = f"{e.completeness * 100:.0f}%"

            missing = []
            if not e.has_check_report:
                missing.append("缺校核报告")
            if not e.handover_package:
                missing.append("缺交接包")
            if not e.has_feedback_report:
                missing.append("缺整改报告")
            note = ", ".join(missing) if missing else "齐全"

            lines.append(
                f"  {d:<12}  {check_icon:<4}  {handover_icon:<4}  "
                f"{fb_icon:<4}  {hist_icon:<4}  {comp:<8}  {note}"
            )

        total = len(dates)
        complete = sum(1 for d in dates if archive[proj][d].completeness >= 0.75)
        missing_days = [d for d in dates if archive[proj][d].completeness < 0.75]

        lines.append("")
        lines.append(f"  项目统计: 共 {total} 天，资料 ≥75% 完整: {complete} 天")
        if missing_days:
            lines.append(f"  需补资料: {', '.join(missing_days[:10])}")
            if len(missing_days) > 10:
                lines.append(f"           等共 {len(missing_days)} 天")
        lines.append("")

    if not projects:
        lines.append("  未找到任何归档资料")
        lines.append("")

    lines.append("=" * 78)
    lines.append("  说明:")
    lines.append("    ✓ 表示该文件存在，✗ 表示缺失")
    lines.append("    完整度 = (有校核报告 + 有交接包 + 有整改报告 + 有复查报告) / 4")
    lines.append("    建议: 完整度 <75% 的日期需要补充资料")
    lines.append("=" * 78)
    lines.append("")
    return "\n".join(lines)


def generate_archive_excel(archive: Dict[str, Dict[str, ArchiveEntry]],
                           out_path: str,
                           project_name: Optional[str] = None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    ws = wb.active
    ws.title = "归档台账"
    ws.append(["桩基施工记录 — 归档台账"])
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center
    ws.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", "", "", "", ""])
    if project_name:
        ws.append(["项目过滤", project_name, "", "", "", "", "", ""])

    headers = [
        "项目名称", "日期", "校核报告", "交接包", "整改报告",
        "复查报告", "完整度", "备注"
    ]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    if project_name:
        projects = [project_name] if project_name in archive else []
    else:
        projects = sorted(archive.keys())

    for proj in projects:
        dates = sorted(archive[proj].keys())
        for d in dates:
            e = archive[proj][d]
            check_val = "✓" if e.has_check_report else "✗"
            handover_val = "✓" if e.handover_package else "✗"
            fb_val = "✓" if e.has_feedback_report else "✗"
            hist_val = "✓" if e.history_report else "✗"
            comp = f"{e.completeness * 100:.0f}%"

            missing = []
            if not e.has_check_report:
                missing.append("缺校核报告")
            if not e.handover_package:
                missing.append("缺交接包")
            if not e.has_feedback_report:
                missing.append("缺整改报告")
            note = ", ".join(missing) if missing else "齐全"

            ws.append([proj, d, check_val, handover_val, fb_val, hist_val, comp, note])
            row = ws.max_row
            for cell in ws[row]:
                cell.border = border

            for col_idx in [3, 4, 5, 6]:
                if ws.cell(row=row, column=col_idx).value == "✓":
                    ws.cell(row=row, column=col_idx).fill = green_fill
                else:
                    ws.cell(row=row, column=col_idx).fill = red_fill

            comp_val = e.completeness
            if comp_val >= 0.75:
                ws.cell(row=row, column=7).fill = green_fill
            elif comp_val >= 0.5:
                ws.cell(row=row, column=7).fill = yellow_fill
            else:
                ws.cell(row=row, column=7).fill = red_fill

            for col_idx in [1, 2, 7, 8]:
                ws.cell(row=row, column=col_idx).alignment = left
            for col_idx in [3, 4, 5, 6]:
                ws.cell(row=row, column=col_idx).alignment = center

    for col, w in zip("ABCDEFGH", [20, 12, 10, 10, 10, 10, 10, 30]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    if not projects:
        ws.append(["未找到任何归档资料"])
        ws.merge_cells(start_row=ws.max_row, start_column=1,
                       end_row=ws.max_row, end_column=8)

    wb.save(out_path)


def save_archive_reports(project_name: Optional[str],
                         output_dir: str,
                         archive: Dict[str, Dict[str, ArchiveEntry]],
                         no_xlsx: bool = False) -> Dict[str, str]:
    os.makedirs(output_dir, exist_ok=True)
    proj_slug = project_name or "all_projects"
    base_name = f"{proj_slug}_归档台账"

    results = {}

    txt = generate_archive_report_text(output_dir, archive, project_name)
    txt_path = os.path.join(output_dir, f"{base_name}.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(txt)
    results["txt"] = txt_path

    if not no_xlsx:
        xlsx_path = os.path.join(output_dir, f"{base_name}.xlsx")
        try:
            generate_archive_excel(archive, xlsx_path, project_name)
            results["xlsx"] = xlsx_path
        except Exception as e:
            print(f"  ⚠ 生成 Excel 归档台账失败: {e}")

    return results
