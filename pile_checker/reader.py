import os
import math
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, field

from .config import CheckConfig, DateFilter, extract_date_from_filename


@dataclass
class PileRecord:
    pile_no: str
    design_length: Optional[float] = None
    actual_hole_depth: Optional[float] = None
    concrete_volume: Optional[float] = None
    theoretical_volume: Optional[float] = None
    hole_finish_time: Optional[datetime] = None
    pouring_start_time: Optional[datetime] = None
    pile_diameter: Optional[float] = None
    source_file: str = ""
    row_index: int = 0

    def calc_theoretical_volume(self, default_diameter: Optional[float] = None) -> Optional[float]:
        if self.theoretical_volume is not None:
            return self.theoretical_volume
        diameter = self.pile_diameter if self.pile_diameter else default_diameter
        if diameter and self.design_length:
            radius = diameter / 2 / 1000.0
            area = math.pi * radius * radius
            return area * self.design_length
        return None

    def record_date(self) -> Optional[datetime]:
        for t in (self.hole_finish_time, self.pouring_start_time):
            if t is not None:
                return datetime(t.year, t.month, t.day)
        return None

    def has_any_data(self) -> bool:
        return bool(
            self.pile_no
            or self.design_length is not None
            or self.actual_hole_depth is not None
            or self.concrete_volume is not None
            or self.theoretical_volume is not None
            or self.pile_diameter is not None
            or self.hole_finish_time is not None
            or self.pouring_start_time is not None
        )


@dataclass
class FileMeta:
    path: str
    name: str
    date_from_name: Optional[datetime] = None
    included: bool = True
    exclude_reason: str = ""
    record_count: int = 0
    total_rows_read: int = 0
    rows_filtered_by_date: int = 0


def _find_column(columns: List[str], candidates: List[str]) -> Optional[str]:
    col_lower = {c.strip().lower(): c for c in columns}
    for cand in candidates:
        cand_lower = cand.strip().lower()
        if cand_lower in col_lower:
            return col_lower[cand_lower]
    return None


def _parse_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_datetime(value) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    try:
        import pandas as pd
        if pd.isna(value):
            return None
    except Exception:
        pass
    s = str(value).strip()
    if not s:
        return None
    fmts = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%Y %H:%M",
        "%m-%d-%Y %H:%M",
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    try:
        import pandas as pd
        ts = pd.to_datetime(s, errors="coerce")
        if pd.notna(ts):
            return ts.to_pydatetime()
    except Exception:
        pass
    return None


def read_csv_file(file_path: str, config: CheckConfig) -> List[PileRecord]:
    import csv
    records = []
    filename = os.path.basename(file_path)
    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return records
        columns = list(reader.fieldnames)
        col_map = {}
        for key in config.field_mapping.keys():
            col = _find_column(columns, config.get_field_names(key))
            if col:
                col_map[key] = col
        for i, row in enumerate(reader, start=2):
            pile_no_raw = row.get(col_map.get("pile_no", ""), "")
            pile_no = str(pile_no_raw).strip() if pile_no_raw is not None else ""
            rec = PileRecord(
                pile_no=pile_no,
                source_file=filename,
                row_index=i,
            )
            if "design_length" in col_map:
                rec.design_length = _parse_float(row.get(col_map["design_length"]))
            if "actual_hole_depth" in col_map:
                rec.actual_hole_depth = _parse_float(row.get(col_map["actual_hole_depth"]))
            if "concrete_volume" in col_map:
                rec.concrete_volume = _parse_float(row.get(col_map["concrete_volume"]))
            if "theoretical_volume" in col_map:
                rec.theoretical_volume = _parse_float(row.get(col_map["theoretical_volume"]))
            if "pile_diameter" in col_map:
                rec.pile_diameter = _parse_float(row.get(col_map["pile_diameter"]))
            if "hole_finish_time" in col_map:
                rec.hole_finish_time = _parse_datetime(row.get(col_map["hole_finish_time"]))
            if "pouring_start_time" in col_map:
                rec.pouring_start_time = _parse_datetime(row.get(col_map["pouring_start_time"]))
            records.append(rec)
    return records


def read_excel_file(file_path: str, config: CheckConfig) -> List[PileRecord]:
    import openpyxl
    records = []
    filename = os.path.basename(file_path)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        header_row = []
        for cell in ws[1]:
            header_row.append(str(cell.value).strip() if cell.value is not None else "")
        col_map = {}
        for key in config.field_mapping.keys():
            col = _find_column(header_row, config.get_field_names(key))
            if col:
                col_idx = header_row.index(col)
                col_map[key] = col_idx
        has_pile_no_col = "pile_no" in col_map
        for row_idx in range(2, ws.max_row + 1):
            row_data = []
            for col_idx in range(ws.max_column):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                row_data.append(cell.value)
            pile_no = ""
            if has_pile_no_col and col_map["pile_no"] < len(row_data):
                pile_no_val = row_data[col_map["pile_no"]]
                pile_no = str(pile_no_val).strip() if pile_no_val is not None else ""
            rec = PileRecord(
                pile_no=pile_no,
                source_file=filename,
                row_index=row_idx,
            )
            if "design_length" in col_map and col_map["design_length"] < len(row_data):
                rec.design_length = _parse_float(row_data[col_map["design_length"]])
            if "actual_hole_depth" in col_map and col_map["actual_hole_depth"] < len(row_data):
                rec.actual_hole_depth = _parse_float(row_data[col_map["actual_hole_depth"]])
            if "concrete_volume" in col_map and col_map["concrete_volume"] < len(row_data):
                rec.concrete_volume = _parse_float(row_data[col_map["concrete_volume"]])
            if "theoretical_volume" in col_map and col_map["theoretical_volume"] < len(row_data):
                rec.theoretical_volume = _parse_float(row_data[col_map["theoretical_volume"]])
            if "pile_diameter" in col_map and col_map["pile_diameter"] < len(row_data):
                rec.pile_diameter = _parse_float(row_data[col_map["pile_diameter"]])
            if "hole_finish_time" in col_map and col_map["hole_finish_time"] < len(row_data):
                rec.hole_finish_time = _parse_datetime(row_data[col_map["hole_finish_time"]])
            if "pouring_start_time" in col_map and col_map["pouring_start_time"] < len(row_data):
                rec.pouring_start_time = _parse_datetime(row_data[col_map["pouring_start_time"]])
            records.append(rec)
    wb.close()
    return records


def filter_records_by_date(records: List[PileRecord], date_filter: DateFilter,
                           file_date: Optional[datetime]) -> Tuple[List[PileRecord], int]:
    if not date_filter.enabled or not date_filter.target_date:
        return records, 0

    target = date_filter.parse_target()
    if target is None:
        return records, 0
    target_day = datetime(target.year, target.month, target.day)

    included = []
    for rec in records:
        rec_day = rec.record_date()
        if rec_day is not None and datetime(rec_day.year, rec_day.month, rec_day.day) == target_day:
            included.append(rec)

    filtered_count = len(records) - len(included)
    return included, filtered_count


def read_all_records(input_dir: str, config: CheckConfig,
                     date_filter: Optional[DateFilter] = None) -> Tuple[List[PileRecord], List[str], List[FileMeta]]:
    all_records: List[PileRecord] = []
    errors: List[str] = []
    file_metas: List[FileMeta] = []

    if not os.path.isdir(input_dir):
        errors.append(f"输入目录不存在: {input_dir}")
        return all_records, errors, file_metas

    supported_ext = (".xlsx", ".xls", ".csv")
    files = []
    for fname in sorted(os.listdir(input_dir)):
        fpath = os.path.join(input_dir, fname)
        if os.path.isfile(fpath) and fname.lower().endswith(supported_ext):
            if not fname.startswith("~$"):
                files.append(fpath)

    if not files:
        errors.append(f"目录 {input_dir} 中未找到 Excel 或 CSV 文件")
        return all_records, errors, file_metas

    for fpath in files:
        fname = os.path.basename(fpath)
        meta = FileMeta(path=fpath, name=fname, date_from_name=extract_date_from_filename(fname))
        try:
            if fpath.lower().endswith(".csv"):
                recs = read_csv_file(fpath, config)
            else:
                recs = read_excel_file(fpath, config)
        except Exception as e:
            errors.append(f"读取文件 {fname} 失败: {str(e)}")
            meta.included = False
            meta.exclude_reason = f"读取失败: {str(e)}"
            file_metas.append(meta)
            continue

        recs = [r for r in recs if r.has_any_data()]
        meta.total_rows_read = len(recs)

        if date_filter and date_filter.enabled:
            recs, filtered_count = filter_records_by_date(recs, date_filter, meta.date_from_name)
            meta.rows_filtered_by_date = filtered_count
            if not recs:
                meta.included = False
                meta.exclude_reason = f"文件中 {meta.total_rows_read} 条记录均不匹配目标日期 {date_filter.target_date}"
                meta.record_count = 0
                file_metas.append(meta)
                continue

        meta.included = True
        meta.record_count = len(recs)
        file_metas.append(meta)
        all_records.extend(recs)

    return all_records, errors, file_metas
