import os
import csv
import io
import json
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass

from .checker import Issue, ISSUE_TYPES


FEEDBACK_STATUS_MAP = {
    "已整改": "已整改",
    "整改完成": "已整改",
    "已核实": "已整改",
    "待复核": "待复核",
    "复核中": "待复核",
    "仍异常": "仍异常",
    "未整改": "仍异常",
    "问题依旧": "仍异常",
    "已关闭": "已整改",
    "关闭": "已整改",
}

STANDARD_STATUSES = ["已整改", "待复核", "仍异常"]


@dataclass
class FeedbackEntry:
    issue_id: str
    status: str
    remark: str = ""
    responder: str = ""
    feedback_date: str = ""
    review_date: str = ""
    handler: str = ""
    attachment: str = ""


def _find_col(fields, candidates):
    fields_lower = {c.strip().lower(): c for c in fields}
    for c in candidates:
        if c.strip().lower() in fields_lower:
            return fields_lower[c.strip().lower()]
    return None


def read_feedback_csv(file_path: str) -> Tuple[List[FeedbackEntry], List[str]]:
    entries = []
    errors = []
    if not os.path.isfile(file_path):
        errors.append(f"反馈文件不存在: {file_path}")
        return entries, errors

    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            errors.append(f"反馈文件为空: {file_path}")
            return entries, errors

        cols = {
            "id": ["问题编号", "issue_id", "编号", "ID"],
            "status": ["整改状态", "状态", "反馈状态", "status"],
            "remark": ["整改说明", "备注", "说明", "remark", "反馈说明"],
            "responder": ["整改人", "反馈人", "负责人", "responder"],
            "feedback_date": ["整改日期", "反馈日期", "日期", "date"],
            "review_date": ["复核日期", "复验日期", "review_date", "reviewDate"],
            "handler": ["处理责任人", "责任人", "处理人", "handler"],
            "attachment": ["附件路径", "附件", "凭证", "attachment"],
        }

        found = {}
        for key, candidates in cols.items():
            found[key] = _find_col(reader.fieldnames, candidates)

        if not found["id"]:
            errors.append(f"反馈文件缺少问题编号列（期望: {', '.join(cols['id'])}）")
            return entries, errors
        if not found["status"]:
            errors.append(f"反馈文件缺少整改状态列（期望: {', '.join(cols['status'])}）")
            return entries, errors

        for i, row in enumerate(reader, start=2):
            def gv(key):
                col = found.get(key)
                if not col:
                    return ""
                return str(row.get(col, "")).strip()

            raw_id = gv("id")
            raw_status = gv("status")
            if not raw_id or not raw_status:
                continue
            normalized = FEEDBACK_STATUS_MAP.get(raw_status, raw_status)
            if normalized not in STANDARD_STATUSES:
                normalized = "待复核"

            entry = FeedbackEntry(
                issue_id=raw_id,
                status=normalized,
                remark=gv("remark"),
                responder=gv("responder"),
                feedback_date=gv("feedback_date"),
                review_date=gv("review_date"),
                handler=gv("handler"),
                attachment=gv("attachment"),
            )
            entries.append(entry)

    return entries, errors


def read_feedback_excel(file_path: str) -> Tuple[List[FeedbackEntry], List[str]]:
    import openpyxl
    entries = []
    errors = []

    if not os.path.isfile(file_path):
        errors.append(f"反馈文件不存在: {file_path}")
        return entries, errors

    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        header = []
        for cell in ws[1]:
            header.append(str(cell.value).strip() if cell.value is not None else "")

        cols = {
            "id": ["问题编号", "issue_id", "编号", "ID"],
            "status": ["整改状态", "状态", "反馈状态", "status"],
            "remark": ["整改说明", "备注", "说明", "remark", "反馈说明"],
            "responder": ["整改人", "反馈人", "负责人", "responder"],
            "feedback_date": ["整改日期", "反馈日期", "日期", "date"],
            "review_date": ["复核日期", "复验日期", "review_date", "reviewDate"],
            "handler": ["处理责任人", "责任人", "处理人", "handler"],
            "attachment": ["附件路径", "附件", "凭证", "attachment"],
        }

        found_idx = {}
        for key, candidates in cols.items():
            col = _find_col(header, candidates)
            found_idx[key] = header.index(col) if col else -1

        if found_idx["id"] < 0 or found_idx["status"] < 0:
            continue

        for row_idx in range(2, ws.max_row + 1):
            def gv(key):
                idx = found_idx.get(key, -1)
                if idx < 0:
                    return ""
                c = ws.cell(row=row_idx, column=idx + 1)
                return str(c.value).strip() if c.value is not None else ""

            raw_id = gv("id")
            raw_status = gv("status")
            if not raw_id or not raw_status:
                continue
            normalized = FEEDBACK_STATUS_MAP.get(raw_status, raw_status)
            if normalized not in STANDARD_STATUSES:
                normalized = "待复核"

            entries.append(FeedbackEntry(
                issue_id=raw_id,
                status=normalized,
                remark=gv("remark"),
                responder=gv("responder"),
                feedback_date=gv("feedback_date"),
                review_date=gv("review_date"),
                handler=gv("handler"),
                attachment=gv("attachment"),
            ))
    wb.close()
    return entries, errors


def read_feedback(file_path: str) -> Tuple[List[FeedbackEntry], List[str]]:
    if file_path.lower().endswith(".csv"):
        return read_feedback_csv(file_path)
    elif file_path.lower().endswith((".xlsx", ".xls")):
        return read_feedback_excel(file_path)
    else:
        return [], [f"不支持的反馈文件格式: {file_path}"]


def merge_feedback(issues: List[Issue], feedback_entries: List[FeedbackEntry]) -> List[Dict]:
    fb_map: Dict[str, FeedbackEntry] = {}
    for entry in feedback_entries:
        fb_map[entry.issue_id] = entry

    results = []
    for issue in issues:
        fb = fb_map.get(issue.issue_id)
        result = {
            "issue_id": issue.issue_id,
            "severity": issue.severity,
            "issue_type": issue.type_label,
            "issue_type_key": issue.issue_type,
            "file_name": issue.file_name,
            "pile_no": issue.pile_no,
            "row_index": issue.row_index,
            "reason": issue.reason,
            "suggestion": issue.suggestion,
            "detail": issue.detail,
            "status": fb.status if fb else "未反馈",
            "remark": fb.remark if fb else "",
            "responder": fb.responder if fb else "",
            "feedback_date": fb.feedback_date if fb else "",
            "review_date": fb.review_date if fb else "",
            "handler": fb.handler if fb else "",
            "attachment": fb.attachment if fb else "",
        }
        results.append(result)
    return results


def status_summary(merged: List[Dict]) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for item in merged:
        st = item["status"]
        counts[st] = counts.get(st, 0) + 1
    return dict(sorted(counts.items()))


def generate_feedback_report_text(project_name: str, check_date: str,
                                  merged: List[Dict]) -> str:
    lines = []
    lines.append("=" * 78)
    lines.append("  桩基施工记录校核 — 整改跟踪汇总报告")
    lines.append("=" * 78)
    lines.append(f"  项目名称: {project_name}")
    lines.append(f"  检查日期: {check_date}")
    lines.append(f"  生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"  问题总数: {len(merged)} 条")
    lines.append("-" * 78)

    summary = status_summary(merged)
    lines.append("")
    lines.append("【整改状态汇总】")
    lines.append("-" * 78)
    for status, count in summary.items():
        icon = {"已整改": "✅", "待复核": "🔄", "仍异常": "❌", "未反馈": "⬜"}.get(status, "❓")
        lines.append(f"  {icon} {status}: {count} 条")

    for group_status in ["仍异常", "待复核", "未反馈", "已整改"]:
        group_items = [m for m in merged if m["status"] == group_status]
        if not group_items:
            continue
        icon = {"已整改": "✅", "待复核": "🔄", "仍异常": "❌", "未反馈": "⬜"}.get(group_status, "❓")
        lines.append("")
        lines.append(f"  ━━━ {icon} {group_status} ━━━")
        lines.append("-" * 78)
        for item in group_items:
            sev_icon = {"严重": "🔴", "警告": "🟡", "提示": "🔵"}.get(item["severity"], "")
            lines.append(
                f"  [{item['issue_id']}] [{item['severity']}] {item['issue_type']} | "
                f"{item['file_name']} 第{item['row_index']}行 | 桩号: {item['pile_no']}"
            )
            lines.append(f"    原因: {item['reason']}")
            if item["remark"]:
                handler_str = f"（{item['handler']}）" if item["handler"] else (
                    f"（{item['responder']}）" if item["responder"] else ""
                )
                lines.append(f"    反馈: {item['remark']}{handler_str}")
            if item["review_date"]:
                lines.append(f"    复核日期: {item['review_date']}")
            if item["feedback_date"]:
                lines.append(f"    反馈日期: {item['feedback_date']}")
            if item["attachment"]:
                lines.append(f"    附件: {item['attachment']}")
            lines.append("")

    lines.append("=" * 78)
    lines.append("  报告结束")
    lines.append("=" * 78)
    lines.append("")
    return "\n".join(lines)


def generate_feedback_excel(merged: List[Dict], out_path: str,
                            project_name: str = "", check_date: str = ""):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = "整改汇总"
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws_summary.append(["桩基校核整改跟踪汇总"])
    ws_summary.merge_cells("A1:D1")
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A1"].alignment = center
    ws_summary.append(["项目名称", project_name, "检查日期", check_date])
    ws_summary.append(["问题总数", len(merged), "生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    ws_summary.append([])
    ws_summary.append(["状态统计"])
    ws_summary.merge_cells(start_row=ws_summary.max_row, start_column=1,
                           end_row=ws_summary.max_row, end_column=2)
    ws_summary.append(["状态", "数量"])
    for cell in ws_summary[ws_summary.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    summary = status_summary(merged)
    status_colors = {
        "已整改": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "待复核": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        "仍异常": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
        "未反馈": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    }
    sev_color = {
        "严重": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
        "警告": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        "提示": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    }
    for status, count in summary.items():
        ws_summary.append([status, count])
        for cell in ws_summary[ws_summary.max_row]:
            cell.border = border
            cell.alignment = center
        color = status_colors.get(status)
        if color:
            ws_summary.cell(row=ws_summary.max_row, column=1).fill = color

    for col, w in zip("ABCD", [16, 16, 16, 22]):
        ws_summary.column_dimensions[col].width = w

    ws_detail = wb.create_sheet("整改明细表")
    headers = [
        "问题编号", "整改状态", "严重程度", "问题类型", "文件名",
        "桩号", "行号", "疑似原因", "建议复核项", "详细信息",
        "整改说明", "整改人", "整改日期", "复核日期", "处理责任人", "附件路径"
    ]
    ws_detail.append(headers)
    for cell in ws_detail[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for item in merged:
        row = [
            item["issue_id"],
            item["status"],
            item["severity"],
            item["issue_type"],
            item["file_name"],
            item["pile_no"],
            item["row_index"],
            item["reason"],
            item["suggestion"],
            item["detail"],
            item["remark"],
            item["responder"],
            item["feedback_date"],
            item["review_date"],
            item["handler"],
            item["attachment"],
        ]
        ws_detail.append(row)
        st_color = status_colors.get(item["status"])
        sv_color = sev_color.get(item["severity"])
        for col_idx, cell in enumerate(ws_detail[ws_detail.max_row], 1):
            cell.border = border
            if col_idx == 2 and st_color:
                cell.fill = st_color
            if col_idx == 3 and sv_color:
                cell.fill = sv_color
            if col_idx in (1, 2, 3, 4, 6, 7):
                cell.alignment = center
            else:
                cell.alignment = left

    widths = [14, 10, 10, 18, 28, 12, 8, 50, 40, 30, 30, 12, 14, 14, 14, 30]
    for idx, w in enumerate(widths, 1):
        ws_detail.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
    ws_detail.freeze_panes = "A2"

    wb.save(out_path)


def save_feedback_reports(project_name: str, check_date: str, output_dir: str,
                          merged: List[Dict], no_xlsx: bool = False) -> Dict[str, str]:
    os.makedirs(output_dir, exist_ok=True)
    safe_date = check_date.replace("/", "-").replace(":", "-")
    base_name = f"{project_name}_{safe_date}_整改跟踪报告"

    results = {}

    txt = generate_feedback_report_text(project_name, check_date, merged)
    txt_path = os.path.join(output_dir, f"{base_name}.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(txt)
    results["txt"] = txt_path

    if not no_xlsx:
        xlsx_path = os.path.join(output_dir, f"{base_name}.xlsx")
        try:
            generate_feedback_excel(merged, xlsx_path, project_name, check_date)
            results["xlsx"] = xlsx_path
        except Exception as e:
            print(f"  ⚠ 生成 Excel 整改报告失败: {e}")

    return results


def generate_feedback_template_csv(project_name: str = "", check_date: str = "") -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "问题编号", "整改状态", "严重程度", "问题类型", "文件名",
        "桩号", "行号", "疑似原因", "建议复核项",
        "整改说明", "整改人", "整改日期", "复核日期", "处理责任人", "附件路径"
    ])
    writer.writerow([
        "示例：PIL-XXXXXXXX", "已整改/待复核/仍异常", "", "", "",
        "", "", "", "",
        "填写整改说明", "施工员姓名", "YYYY-MM-DD", "YYYY-MM-DD", "责任人姓名", "照片或文档路径"
    ])
    return output.getvalue()
