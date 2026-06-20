import os
import csv
import io
from datetime import datetime
from typing import List, Dict, Optional

from .checker import Issue, ISSUE_TYPES
from .config import ISSUE_SEVERITY, SEVERITY_ORDER
from .reader import FileMeta


SEVERITY_ICONS = {"严重": "🔴", "警告": "🟡", "提示": "🔵"}


def _severity_counts(issues: List[Issue]) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for issue in issues:
        sev = issue.severity
        counts[sev] = counts.get(sev, 0) + 1
    return dict(sorted(counts.items(), key=lambda x: SEVERITY_ORDER.get(x[0], 99)))


def _type_counts(issues: List[Issue]) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for issue in issues:
        counts[issue.issue_type] = counts.get(issue.issue_type, 0) + 1
    return dict(sorted(counts.items(), key=lambda x: -x[1]))


def generate_text_report(project_name: str, check_date: str, issues: List[Issue],
                         total_records: int, total_files: int,
                         errors: List[str], file_metas: Optional[List[FileMeta]] = None) -> str:
    lines = []
    lines.append("=" * 78)
    lines.append("  桩基施工记录批量校核报告")
    lines.append("=" * 78)
    lines.append(f"  项目名称: {project_name}")
    lines.append(f"  检查日期: {check_date}")
    lines.append(f"  生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"  纳入检查文件: {total_files} 份")
    lines.append(f"  有效记录总数: {total_records} 条")
    lines.append(f"  问题合计: {len(issues)} 条")
    lines.append("-" * 78)

    if file_metas:
        included = [m for m in file_metas if m.included]
        excluded = [m for m in file_metas if not m.included]
        total_filtered = sum(m.rows_filtered_by_date for m in file_metas)
        if included:
            lines.append("")
            lines.append("【本次纳入文件】")
            lines.append("-" * 78)
            for m in included:
                date_str = m.date_from_name.strftime("%Y-%m-%d") if m.date_from_name else "未识别"
                filter_note = ""
                if m.rows_filtered_by_date > 0:
                    filter_note = f"  [筛除非当天 {m.rows_filtered_by_date} 条]"
                lines.append(f"  · {m.name}  [{date_str}]  纳入 {m.record_count} 条{filter_note}")
        if total_filtered > 0:
            lines.append("")
            lines.append(f"  ※ 日期筛选共排除非当天记录 {total_filtered} 条（按记录行内日期逐行筛选）")
        if excluded:
            lines.append("")
            lines.append("【已排除文件】")
            lines.append("-" * 78)
            for m in excluded:
                date_str = m.date_from_name.strftime("%Y-%m-%d") if m.date_from_name else "未识别"
                lines.append(f"  ✗ {m.name}  [{date_str}]  → {m.exclude_reason}")

    if errors:
        lines.append("")
        lines.append("【读取错误】")
        lines.append("-" * 78)
        for i, err in enumerate(errors, 1):
            lines.append(f"  {i}. {err}")

    if not issues:
        lines.append("")
        lines.append("【检查结果】全部通过，未发现明显问题。")
        lines.append("  （注：本工具仅作批量初筛，正式资料仍需人工抽核。）")
        lines.append("")
        lines.append("=" * 78)
        lines.append("  报告结束")
        lines.append("=" * 78)
        lines.append("")
        return "\n".join(lines)

    sev_counts = _severity_counts(issues)
    type_counts = _type_counts(issues)

    lines.append("")
    lines.append("【按严重程度统计】")
    lines.append("-" * 78)
    for sev, count in sev_counts.items():
        icon = SEVERITY_ICONS.get(sev, "")
        lines.append(f"  {icon} {sev}: {count} 条")

    lines.append("")
    lines.append("【按问题类型统计】")
    lines.append("-" * 78)
    for itype, count in type_counts.items():
        label = ISSUE_TYPES.get(itype, itype)
        sev = ISSUE_SEVERITY.get(itype, "提示")
        lines.append(f"  {label}（{sev}）: {count} 条")

    lines.append("")
    lines.append("【问题明细】按严重程度 → 问题类型 → 文件 → 行号 排序")
    lines.append("-" * 78)

    current_severity = None
    current_type = None
    current_file = None
    idx = 1

    for issue in issues:
        if issue.severity != current_severity:
            current_severity = issue.severity
            current_type = None
            current_file = None
            icon = SEVERITY_ICONS.get(current_severity, "")
            lines.append("")
            lines.append(f"  ━━━ {icon} {current_severity}问题 ━━━")

        if issue.issue_type != current_type:
            current_type = issue.issue_type
            current_file = None
            lines.append(f"    ◆ {issue.type_label}")

        if issue.file_name != current_file:
            current_file = issue.file_name
            lines.append(f"      ▶ {current_file}")

        sev_icon = SEVERITY_ICONS.get(issue.severity, "")
        lines.append(f"      {sev_icon} [{idx}] ({issue.issue_id}) 桩号: {issue.pile_no}")
        lines.append(f"           疑似原因: {issue.reason}")
        lines.append(f"           建议复核: {issue.suggestion}")
        if issue.detail:
            lines.append(f"           定位信息: {issue.detail}")
        idx += 1

    lines.append("")
    lines.append("=" * 78)
    lines.append("  报告结束  请内业工程师逐条核实后反馈施工员整改")
    lines.append("=" * 78)
    lines.append("")
    return "\n".join(lines)


def generate_csv_report(issues: List[Issue], project_name: str = "",
                        check_date: str = "") -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "问题编号", "序号", "严重程度", "问题类型", "项目名称", "检查日期",
        "文件名", "桩号", "行号", "疑似原因", "建议复核项",
        "详细信息", "是否整改", "整改说明"
    ])
    for i, issue in enumerate(issues, 1):
        writer.writerow([
            issue.issue_id,
            i,
            issue.severity,
            issue.type_label,
            project_name,
            check_date,
            issue.file_name,
            issue.pile_no,
            issue.row_index if issue.row_index else "",
            issue.reason,
            issue.suggestion,
            issue.detail,
            "",
            "",
        ])
    return output.getvalue()


def generate_excel_report(issues: List[Issue], out_path: str,
                          project_name: str = "", check_date: str = ""):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = "汇总"
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    sev_counts = _severity_counts(issues)
    type_counts = _type_counts(issues)

    ws_summary.append(["桩基施工记录校核报告 - 汇总页"])
    ws_summary.merge_cells("A1:F1")
    c = ws_summary["A1"]
    c.font = Font(bold=True, size=14)
    c.alignment = center

    ws_summary.append(["项目名称", project_name, "", "检查日期", check_date, ""])
    ws_summary.append(["问题总数", len(issues), "", "生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""])

    ws_summary.append([])
    ws_summary.append(["按严重程度统计"])
    ws_summary.merge_cells(start_row=ws_summary.max_row, start_column=1,
                           end_row=ws_summary.max_row, end_column=2)
    ws_summary.append(["严重程度", "数量"])
    for cell in ws_summary[ws_summary.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    for sev, count in sev_counts.items():
        ws_summary.append([sev, count])
        for cell in ws_summary[ws_summary.max_row]:
            cell.border = border
            cell.alignment = center

    ws_summary.append([])
    ws_summary.append(["按问题类型统计"])
    ws_summary.merge_cells(start_row=ws_summary.max_row, start_column=1,
                           end_row=ws_summary.max_row, end_column=3)
    ws_summary.append(["问题类型", "严重程度", "数量"])
    for cell in ws_summary[ws_summary.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    for itype, count in type_counts.items():
        label = ISSUE_TYPES.get(itype, itype)
        sev = ISSUE_SEVERITY.get(itype, "提示")
        ws_summary.append([label, sev, count])
        for cell in ws_summary[ws_summary.max_row]:
            cell.border = border
            cell.alignment = center

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 22
    ws_summary.column_dimensions["C"].width = 12
    ws_summary.column_dimensions["D"].width = 16
    ws_summary.column_dimensions["E"].width = 22
    ws_summary.column_dimensions["F"].width = 12

    ws_detail = wb.create_sheet("问题明细表")
    headers = [
        "问题编号", "序号", "严重程度", "问题类型", "文件名", "桩号", "行号",
        "疑似原因", "建议复核项", "详细信息", "是否整改", "整改说明"
    ]
    ws_detail.append(headers)
    for cell in ws_detail[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    sev_color = {
        "严重": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
        "警告": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        "提示": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    }

    for i, issue in enumerate(issues, 1):
        row = [
            issue.issue_id,
            i,
            issue.severity,
            issue.type_label,
            issue.file_name,
            issue.pile_no,
            issue.row_index if issue.row_index else "",
            issue.reason,
            issue.suggestion,
            issue.detail,
            "",
            "",
        ]
        ws_detail.append(row)
        color_fill = sev_color.get(issue.severity)
        for col_idx, cell in enumerate(ws_detail[ws_detail.max_row], 1):
            cell.border = border
            if col_idx == 3 and color_fill:
                cell.fill = color_fill
            if col_idx in (1, 2, 3, 4, 6, 7):
                cell.alignment = center
            else:
                cell.alignment = left

    widths = [14, 6, 10, 18, 28, 12, 8, 50, 40, 30, 10, 30]
    for idx, w in enumerate(widths, 1):
        ws_detail.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
    ws_detail.freeze_panes = "A2"

    wb.save(out_path)


def save_reports(project_name: str, check_date: str, output_dir: str,
                 issues: List[Issue], total_records: int, total_files: int,
                 errors: List[str], file_metas: Optional[List[FileMeta]] = None,
                 no_xlsx: bool = False) -> Dict[str, str]:
    os.makedirs(output_dir, exist_ok=True)
    safe_date = check_date.replace("/", "-").replace(":", "-")
    base_name = f"{project_name}_{safe_date}_桩基校核报告"

    results = {}

    text_content = generate_text_report(project_name, check_date, issues,
                                        total_records, total_files, errors, file_metas)
    txt_path = os.path.join(output_dir, f"{base_name}.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(text_content)
    results["txt"] = txt_path

    if issues:
        csv_content = generate_csv_report(issues, project_name, check_date)
        csv_path = os.path.join(output_dir, f"{base_name}.csv")
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            f.write(csv_content)
        results["csv"] = csv_path

        if not no_xlsx:
            xlsx_path = os.path.join(output_dir, f"{base_name}.xlsx")
            try:
                generate_excel_report(issues, xlsx_path, project_name, check_date)
                results["xlsx"] = xlsx_path
            except Exception as e:
                print(f"  ⚠ 生成 Excel 报告失败: {e}（已回退到 CSV）")

    return results


def print_console_summary(issues: List[Issue], total_records: int,
                          total_files: int, errors: List[str],
                          file_metas: Optional[List[FileMeta]] = None):
    print()
    print("━" * 64)

    if file_metas:
        included = [m for m in file_metas if m.included]
        excluded = [m for m in file_metas if not m.included]
        total_filtered = sum(m.rows_filtered_by_date for m in file_metas)
        print(f"  扫描文件 {len(file_metas)} 份，纳入检查 {len(included)} 份，排除 {len(excluded)} 份")
        if included:
            print(f"  纳入文件:")
            for m in included:
                date_str = m.date_from_name.strftime("%m-%d") if m.date_from_name else "日期未识别"
                filter_note = ""
                if m.rows_filtered_by_date > 0:
                    filter_note = f" (筛除非当天 {m.rows_filtered_by_date} 条)"
                print(f"    · {m.name}  [{date_str}]  纳入 {m.record_count} 条{filter_note}")
        if total_filtered > 0:
            print(f"  ※ 日期筛选共排除非当天记录 {total_filtered} 条")
        if excluded:
            print(f"  排除文件:")
            for m in excluded:
                print(f"    ✗ {m.name}  → {m.exclude_reason}")
    else:
        print(f"  共检查 {total_files} 份文件")

    print(f"  有效记录 {total_records} 条")
    if errors:
        print(f"  读取错误 {len(errors)} 项")
    print(f"  发现问题 {len(issues)} 条")

    if issues:
        sev_counts = _severity_counts(issues)
        print("━" * 64)
        print("  按严重程度:")
        for sev, count in sev_counts.items():
            icon = SEVERITY_ICONS.get(sev, "")
            print(f"    {icon} {sev}: {count} 条")
        print("  按问题类型:")
        type_counts = _type_counts(issues)
        for itype, count in type_counts.items():
            label = ISSUE_TYPES.get(itype, itype)
            sev = ISSUE_SEVERITY.get(itype, "提示")
            print(f"    • [{sev}] {label}: {count} 条")
    else:
        print("━" * 64)
        print("  ✓ 全部通过，未发现明显问题。")
        print("    （建议仍需人工抽核关键记录）")
    print()
