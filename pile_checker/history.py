import os
import csv
import io
import json
import re
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, field


DATE_PATTERNS = [
    r"(\d{4})[-_](\d{1,2})[-_](\d{1,2})",
    r"(\d{4})(\d{2})(\d{2})",
]


@dataclass
class DailyIssueSnapshot:
    date: str
    issue_id: str
    severity: str
    issue_type: str
    file_name: str
    pile_no: str
    reason: str
    status: str = "未反馈"


@dataclass
class DailySummary:
    date: str
    total_issues: int = 0
    new_issues: List[str] = field(default_factory=list)
    persistent_issues: List[str] = field(default_factory=list)
    resolved_issues: List[str] = field(default_factory=list)
    by_severity: Dict[str, int] = field(default_factory=dict)
    by_type: Dict[str, int] = field(default_factory=dict)


def _extract_date_from_filename(filename: str) -> Optional[str]:
    for pat in DATE_PATTERNS:
        m = re.search(pat, filename)
        if m:
            try:
                y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
                return datetime(y, mo, d).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def _find_report_files(report_dir: str, project_name: Optional[str] = None) -> List[Tuple[str, str]]:
    files = []
    if not os.path.isdir(report_dir):
        return files
    for name in sorted(os.listdir(report_dir)):
        if not name.endswith("_桩基校核报告.csv"):
            continue
        if project_name:
            prefix = f"{project_name}_"
            if not name.startswith(prefix):
                continue
        date_str = _extract_date_from_filename(name)
        if date_str:
            files.append((date_str, os.path.join(report_dir, name)))
    return files


def _find_longest_consecutive_streak(dates: List[str]) -> Tuple[int, List[str]]:
    if not dates:
        return 0, []
    sorted_dates = sorted(set(dates))
    max_len = 1
    max_start = 0
    cur_len = 1
    cur_start = 0
    for i in range(1, len(sorted_dates)):
        d_prev = datetime.strptime(sorted_dates[i - 1], "%Y-%m-%d")
        d_curr = datetime.strptime(sorted_dates[i], "%Y-%m-%d")
        if (d_curr - d_prev).days == 1:
            cur_len += 1
            if cur_len > max_len:
                max_len = cur_len
                max_start = cur_start
        else:
            cur_len = 1
            cur_start = i
    streak = sorted_dates[max_start:max_start + max_len]
    return max_len, streak


def load_issues_from_csv(csv_path: str) -> Dict[str, DailyIssueSnapshot]:
    issues = {}
    if not os.path.isfile(csv_path):
        return issues
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return issues
        has_issue_id = "问题编号" in reader.fieldnames
        for row in reader:
            if has_issue_id:
                issue_id = str(row.get("问题编号", "")).strip()
            else:
                raw = f"{row.get('文件名','')}|{row.get('桩号','')}|{row.get('疑似原因','')}"
                import hashlib
                issue_id = f"OLD-{hashlib.md5(raw.encode()).hexdigest()[:8].upper()}"
            if not issue_id:
                continue
            snap = DailyIssueSnapshot(
                date=_extract_date_from_filename(os.path.basename(csv_path)) or "",
                issue_id=issue_id,
                severity=str(row.get("严重程度", "")).strip(),
                issue_type=str(row.get("问题类型", "")).strip(),
                file_name=str(row.get("文件名", "")).strip(),
                pile_no=str(row.get("桩号", "")).strip(),
                reason=str(row.get("疑似原因", "")).strip(),
            )
            issues[issue_id] = snap
    return issues


def build_history(report_dir: str,
                  project_name: Optional[str] = None,
                  start_date: Optional[str] = None,
                  end_date: Optional[str] = None) -> Dict[str, DailySummary]:
    all_files = _find_report_files(report_dir, project_name)
    if not all_files:
        return {}

    if start_date:
        all_files = [(d, p) for d, p in all_files if d >= start_date]
    if end_date:
        all_files = [(d, p) for d, p in all_files if d <= end_date]
    all_files.sort(key=lambda x: x[0])

    history: Dict[str, DailySummary] = {}
    prev_ids: set = set()

    for date_str, csv_path in all_files:
        day_issues = load_issues_from_csv(csv_path)
        day_ids = set(day_issues.keys())

        summary = DailySummary(date=date_str, total_issues=len(day_ids))
        summary.new_issues = sorted(day_ids - prev_ids)
        summary.persistent_issues = sorted(day_ids & prev_ids)
        summary.resolved_issues = sorted(prev_ids - day_ids) if prev_ids else []

        for issue_id in day_ids:
            snap = day_issues[issue_id]
            summary.by_severity[snap.severity] = summary.by_severity.get(snap.severity, 0) + 1
            summary.by_type[snap.issue_type] = summary.by_type.get(snap.issue_type, 0) + 1

        history[date_str] = summary
        prev_ids = day_ids

    return history


def find_persistent_issues(report_dir: str,
                           project_name: Optional[str] = None,
                           min_days: int = 2,
                           start_date: Optional[str] = None,
                           end_date: Optional[str] = None) -> Dict[str, List[str]]:
    all_files = _find_report_files(report_dir, project_name)
    if not all_files:
        return {}, {}
    if start_date:
        all_files = [(d, p) for d, p in all_files if d >= start_date]
    if end_date:
        all_files = [(d, p) for d, p in all_files if d <= end_date]
    all_files.sort(key=lambda x: x[0])

    appearances: Dict[str, List[str]] = {}
    all_details: Dict[str, DailyIssueSnapshot] = {}

    for date_str, csv_path in all_files:
        day_issues = load_issues_from_csv(csv_path)
        for issue_id, snap in day_issues.items():
            appearances.setdefault(issue_id, []).append(date_str)
            if issue_id not in all_details:
                all_details[issue_id] = snap

    persistent = {}
    for issue_id, all_dates in appearances.items():
        streak_len, streak_dates = _find_longest_consecutive_streak(all_dates)
        if streak_len >= min_days:
            persistent[issue_id] = streak_dates

    return persistent, all_details


def generate_history_report_text(project_name: str,
                                  history: Dict[str, DailySummary],
                                  persistent: Optional[Dict[str, List[str]]] = None,
                                  details: Optional[Dict[str, DailyIssueSnapshot]] = None,
                                  start_date: str = "",
                                  end_date: str = "") -> str:
    lines = []
    lines.append("=" * 78)
    lines.append("  桩基施工记录校核 — 多日复查视图")
    lines.append("=" * 78)
    lines.append(f"  项目名称: {project_name}")
    if start_date or end_date:
        lines.append(f"  日期范围: {start_date or '最早'} ~ {end_date or '最新'}")
    lines.append(f"  生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"  覆盖天数: {len(history)} 天")
    lines.append("-" * 78)

    dates = sorted(history.keys())

    lines.append("")
    lines.append("【每日概览】")
    lines.append("-" * 78)
    lines.append(f"  {'日期':<12}  {'问题':>5}  {'新增':>5}  {'持续':>5}  {'已解决':>6}  {'严重':>4}  {'警告':>4}  {'提示':>4}")
    lines.append(f"  {'─' * 12}  {'─'*5}  {'─'*5}  {'─'*5}  {'─'*6}  {'─'*4}  {'─'*4}  {'─'*4}")
    for d in dates:
        s = history[d]
        lines.append(
            f"  {d:<12}  {s.total_issues:>5}  {len(s.new_issues):>5}  "
            f"{len(s.persistent_issues):>5}  {len(s.resolved_issues):>6}  "
            f"{s.by_severity.get('严重', 0):>4}  "
            f"{s.by_severity.get('警告', 0):>4}  "
            f"{s.by_severity.get('提示', 0):>4}"
        )

    for d in dates:
        s = history[d]
        lines.append("")
        lines.append(f"  ▶ {d}  共 {s.total_issues} 条问题")
        if s.new_issues:
            lines.append(f"    新增 ({len(s.new_issues)}): {', '.join(s.new_issues[:8])}")
            if len(s.new_issues) > 8:
                lines.append(f"                 等共 {len(s.new_issues)} 条")
        if s.persistent_issues:
            lines.append(f"    持续 ({len(s.persistent_issues)}): {', '.join(s.persistent_issues[:8])}")
            if len(s.persistent_issues) > 8:
                lines.append(f"                 等共 {len(s.persistent_issues)} 条")
        if s.resolved_issues:
            lines.append(f"    已解决 ({len(s.resolved_issues)}): {', '.join(s.resolved_issues[:8])}")
            if len(s.resolved_issues) > 8:
                lines.append(f"                 等共 {len(s.resolved_issues)} 条")

    if persistent and details:
        lines.append("")
        lines.append("【连续多天未关闭的问题】")
        lines.append("-" * 78)
        sorted_persistent = sorted(persistent.items(), key=lambda x: -len(x[1]))
        for issue_id, date_list in sorted_persistent:
            snap = details.get(issue_id)
            if snap:
                lines.append(
                    f"  [{issue_id}] [{snap.severity}] {snap.issue_type} | "
                    f"连续 {len(date_list)} 天 | {snap.file_name} 桩号:{snap.pile_no}"
                )
                lines.append(f"    出现日期: {', '.join(date_list)}")
                lines.append(f"    原因: {snap.reason}")
                lines.append("")

    lines.append("=" * 78)
    lines.append("  报告结束")
    lines.append("=" * 78)
    lines.append("")
    return "\n".join(lines)


def generate_history_excel(project_name: str, history: Dict[str, DailySummary],
                           out_path: str,
                           persistent: Optional[Dict[str, List[str]]] = None,
                           details: Optional[Dict[str, DailyIssueSnapshot]] = None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sev_color = {
        "严重": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
        "警告": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        "提示": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    }

    ws_daily = wb.active
    ws_daily.title = "每日概览"
    ws_daily.append(["桩基校核 - 多日复查日报"])
    ws_daily.merge_cells("A1:H1")
    ws_daily["A1"].font = Font(bold=True, size=14)
    ws_daily["A1"].alignment = center
    ws_daily.append(["项目名称", project_name, "生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", "", ""])

    headers = ["日期", "问题总数", "新增", "持续", "已解决", "严重", "警告", "提示"]
    ws_daily.append(headers)
    for cell in ws_daily[ws_daily.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for d in sorted(history.keys()):
        s = history[d]
        ws_daily.append([
            d, s.total_issues, len(s.new_issues), len(s.persistent_issues),
            len(s.resolved_issues),
            s.by_severity.get("严重", 0),
            s.by_severity.get("警告", 0),
            s.by_severity.get("提示", 0),
        ])
        for cell in ws_daily[ws_daily.max_row]:
            cell.border = border
            cell.alignment = center

    for col, w in zip("ABCDEFGH", [14, 10, 8, 8, 10, 8, 8, 8]):
        ws_daily.column_dimensions[col].width = w

    ws_detail = wb.create_sheet("每日新增明细")
    ws_detail.append(["日期", "问题编号", "严重程度", "问题类型", "文件名", "桩号", "疑似原因", "分类"])
    for cell in ws_detail[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    all_issue_details: Dict[str, DailyIssueSnapshot] = details or {}

    for d in sorted(history.keys()):
        s = history[d]
        for issue_id in s.new_issues:
            snap = all_issue_details.get(issue_id)
            sev = snap.severity if snap else ""
            itype = snap.issue_type if snap else ""
            fname = snap.file_name if snap else ""
            pno = snap.pile_no if snap else ""
            reason = snap.reason if snap else ""
            ws_detail.append([d, issue_id, sev, itype, fname, pno, reason, "新增"])
            row_num = ws_detail.max_row
            for cell in ws_detail[row_num]:
                cell.border = border
            color = sev_color.get(sev)
            if color:
                ws_detail.cell(row=row_num, column=3).fill = color

        for issue_id in s.persistent_issues:
            snap = all_issue_details.get(issue_id)
            sev = snap.severity if snap else ""
            itype = snap.issue_type if snap else ""
            fname = snap.file_name if snap else ""
            pno = snap.pile_no if snap else ""
            reason = snap.reason if snap else ""
            ws_detail.append([d, issue_id, sev, itype, fname, pno, reason, "持续"])
            row_num = ws_detail.max_row
            for cell in ws_detail[row_num]:
                cell.border = border
            color = sev_color.get(sev)
            if color:
                ws_detail.cell(row=row_num, column=3).fill = color

    widths = [12, 14, 10, 18, 28, 12, 50, 8]
    for idx, w in enumerate(widths, 1):
        ws_detail.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
    ws_detail.freeze_panes = "A2"

    if persistent and details:
        ws_per = wb.create_sheet("连续未关闭")
        ws_per.append(["问题编号", "出现天数", "严重程度", "问题类型", "文件名", "桩号", "出现日期", "疑似原因"])
        for cell in ws_per[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        sorted_p = sorted(persistent.items(), key=lambda x: -len(x[1]))
        for issue_id, date_list in sorted_p:
            snap = details.get(issue_id)
            if snap:
                ws_per.append([
                    issue_id, len(date_list), snap.severity, snap.issue_type,
                    snap.file_name, snap.pile_no, ", ".join(date_list), snap.reason,
                ])
                row_num = ws_per.max_row
                for cell in ws_per[row_num]:
                    cell.border = border
                color = sev_color.get(snap.severity)
                if color:
                    ws_per.cell(row=row_num, column=3).fill = color

        widths = [14, 10, 10, 18, 28, 12, 50, 50]
        for idx, w in enumerate(widths, 1):
            ws_per.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
        ws_per.freeze_panes = "A2"

    wb.save(out_path)


def save_history_reports(project_name: str, output_dir: str,
                         history: Dict[str, DailySummary],
                         persistent: Optional[Dict[str, List[str]]] = None,
                         details: Optional[Dict[str, DailyIssueSnapshot]] = None,
                         start_date: str = "", end_date: str = "",
                         no_xlsx: bool = False) -> Dict[str, str]:
    os.makedirs(output_dir, exist_ok=True)
    suffix = f"{start_date or 'all'}_to_{end_date or 'latest'}"
    base_name = f"{project_name}_{suffix}_多日复查"

    results = {}

    txt = generate_history_report_text(project_name, history, persistent, details, start_date, end_date)
    txt_path = os.path.join(output_dir, f"{base_name}.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(txt)
    results["txt"] = txt_path

    if not no_xlsx:
        xlsx_path = os.path.join(output_dir, f"{base_name}.xlsx")
        try:
            generate_history_excel(project_name, history, xlsx_path, persistent, details)
            results["xlsx"] = xlsx_path
        except Exception as e:
            print(f"  ⚠ 生成 Excel 历史报告失败: {e}")

    return results
